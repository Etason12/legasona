import hmac
import logging
import os
import time
from flask import Blueprint, jsonify, request
from flask_jwt_extended import jwt_required, get_jwt_identity
from app.models import Branch, User, Vehicle, SparePart, Sale, Payment, Order, Transfer, Purchase, PurchaseItem, Expense, ActivityLog, Customer
from app.utils.auth import admin_required

logger = logging.getLogger(__name__)

system_bp = Blueprint('system', __name__)

_DEFAULT_SECRET = 'dev-secret-key'
_DEFAULT_RESET_KEY = 'dev-reset-key'

def _valid_reset_key(reset_key):
    secret = os.environ.get('ADMIN_RESET_KEY', _DEFAULT_RESET_KEY)
    if secret == _DEFAULT_RESET_KEY:
        return False
    return hmac.compare_digest(reset_key, secret)

@system_bp.route('/import-excel-vehicles', methods=['POST'])
@jwt_required()
@admin_required
def import_excel_vehicles():
    from app import db
    import io
    try:
        import openpyxl
        uploaded = request.files.get('file')
        if uploaded and uploaded.filename:
            wb = openpyxl.load_workbook(io.BytesIO(uploaded.read()), data_only=True)
        else:
            # Fallback: read a tracked import file stored in the repo root
            excel_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '../../../attachment.xlsx'))
            if not os.path.exists(excel_path):
                return jsonify({'message': 'No Excel file uploaded and no import file found on the server'}), 400
            wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active
        if ws.max_row < 2:
            return jsonify({'message': 'Excel file has no data rows'}), 400
        headers = [str(c.value).strip() if c.value is not None else '' for c in ws[1]]
        norm = {h.lower(): h for h in headers if h}

        def col(*aliases):
            for alias in aliases:
                if alias in norm:
                    return headers.index(norm[alias])
            return None

        idx = {
            'vin': col('vehicle identification number', 'serial/vin', 'vin'),
            'model': col('model'),
            'engine_number': col('specification', 'motor/engine no', 'engine number'),
            'color': col('color', 'colour'),
            'status': col('operation status', 'status'),
            'type': col('body classification', 'type'),
            'power_type': col('propulsion type', 'propulsion'),
            'selling_price': col('selling price', 'price'),
            'cost_price': col('cost', 'cost price'),
        }
        if idx['vin'] is None or idx['selling_price'] is None:
            return jsonify({'message': 'Required columns "Vehicle Identification Number" and "Selling Price" not found in Excel'}), 400

        branch = Branch.query.filter_by(name='Mekelle').first() or Branch.query.first()
        if not branch:
            return jsonify({'message': 'No branch found in database'}), 400

        def cell(row, key):
            i = idx.get(key)
            return row[i] if i is not None and i < len(row) else None

        imported = 0
        skipped = 0
        seen = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            raw_vin = cell(row, 'vin')
            if raw_vin is None or str(raw_vin).strip() in ('', 'None'):
                continue
            vin = str(raw_vin).strip().upper()
            if vin in seen or Vehicle.query.filter_by(vin=vin).first():
                skipped += 1
                continue
            seen.add(vin)
            raw_model = cell(row, 'model')
            raw_engine = cell(row, 'engine_number')
            raw_color = cell(row, 'color')
            raw_status = cell(row, 'status')
            raw_type = cell(row, 'type')
            raw_power = cell(row, 'power_type')
            raw_price = cell(row, 'selling_price')
            model = str(raw_model).strip() if raw_model not in (None, '') else ''
            engine = str(raw_engine).strip() if raw_engine not in (None, '') else ''
            color = str(raw_color).strip() if raw_color not in (None, '') else ''
            status = (str(raw_status).strip().lower() if raw_status not in (None, '') else 'available') or 'available'
            vtype = (str(raw_type).strip().lower() if raw_type not in (None, '') else '3-wheel') or '3-wheel'
            ptype = (str(raw_power).strip().lower() if raw_power not in (None, '') else 'electric') or 'electric'
            try:
                selling_price = float(raw_price) if raw_price not in (None, '') else 0
            except (TypeError, ValueError):
                selling_price = 0
            raw_cost = cell(row, 'cost_price')
            if raw_cost not in (None, ''):
                try:
                    cost_price = float(raw_cost)
                except (TypeError, ValueError):
                    cost_price = 0
            else:
                # No cost column: assume 70% of the selling price as estimated cost
                cost_price = round(selling_price * 0.7, 2)
            v = Vehicle(
                vin=vin, model=model, engine_number=engine, color=color,
                status=status, type=vtype, power_type=ptype,
                chassis_number=vin, selling_price=selling_price, cost_price=cost_price,
                branch_id=branch.id,
            )
            db.session.add(v)
            imported += 1
        db.session.commit()
        return jsonify({'message': f'Imported {imported} vehicles, skipped {skipped}'}), 200
    except Exception as e:
        db.session.rollback()
        logger.error(f'Import excel vehicles failed: {e}', exc_info=True)
        return jsonify({'message': 'Import failed. Please check the file format and try again.'}), 500

@system_bp.route('/reset-database', methods=['POST'])
@jwt_required()
@admin_required
def reset_database():
    data = request.get_json() or {}
    if not _valid_reset_key(data.get('reset_key', '')):
        return jsonify({'message': 'Invalid reset key'}), 401
    from app import db
    try:
        Payment.query.delete()
        Sale.query.delete()
        Order.query.delete()
        Transfer.query.delete()
        PurchaseItem.query.delete()
        Purchase.query.delete()
        Expense.query.delete()
        ActivityLog.query.delete()
        Customer.query.delete()
        Vehicle.query.delete()
        SparePart.query.delete()
        User.query.delete()
        Branch.query.delete()
        db.session.commit()

        shire = Branch(name='Shire', location='Shire, Tigray')
        mekelle = Branch(name='Mekelle', location='Mekelle, Tigray')
        db.session.add_all([shire, mekelle])
        db.session.flush()

        admin = User(username='admin', role='admin', branch_id=shire.id)
        admin.set_password(data.get('new_password') or 'admin123')
        db.session.add(admin)

        vehicles = [
            Vehicle(vin='HILUX-4WD-001', type='4-wheel', power_type='non-electric', model='Toyota Hilux 4x4 2025', chassis_number='HILUX-4WD-001', engine_number='1KD-FTV-88421', color='White', branch_id=shire.id, status='available', cost_price=3200000, selling_price=4500000),
            Vehicle(vin='FOTON-EV-3W-003', type='3-wheel', power_type='electric', model='Foton Electric Tricycle', chassis_number='FOTON-EV-3W-003', engine_number='MOT-EV-33210', color='Blue', branch_id=mekelle.id, status='available', cost_price=800000, selling_price=1200000),
        ]
        db.session.add_all(vehicles)

        parts = [
            SparePart(part_number='OIL-FILT-001', name='Engine Oil Filter', category='Filters', quantity=45, branch_id=shire.id, unit_price=1200, cost_price=450),
            SparePart(part_number='BRK-PAD-002', name='Brake Pad Set (Front)', category='Brakes', quantity=12, branch_id=mekelle.id, unit_price=4500, cost_price=1800),
        ]
        db.session.add_all(parts)

        db.session.commit()
        return jsonify({'message': 'Database reset successfully. Default admin user and sample data restored.'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'Reset failed'}), 500


@system_bp.route('/import-spare-parts', methods=['POST'])
@jwt_required()
@admin_required
def import_spare_parts():
    from app import db
    PARTS = [
        ('ቅድመት መብራህቲ ከቨር ናይ ፀጋም', 'Left Front Light Cover', 6, 'Lighting & Exterior'),
        ('ቅድመት መብራህቲ ከቨር ናይ የማን', 'Right Front Light Cover', 7, 'Lighting & Exterior'),
        ('ማእኸል ከቨር', 'Center Cover', 6, 'Body & Exterior'),
        ('አንቴና', 'Antenna', 20, 'Audio & Electronics'),
        ('ናይ ድሕሪት ጡሩንባ', 'Rear Horn', 4, 'Horn & Electrical'),
        ('H & L እጀታ', 'High/Low Beam Switch', 4, 'Lighting & Electrical'),
        ('R & D እጀታ', 'R & D Handle / Reverse & Drive Lever', 4, 'Controls & Transmission'),
        ('ስቶፕ ላይት ኬብል', 'Stop Light Cable', 4, 'Lighting & Electrical'),
        ('ክላክስ', 'Horn', 8, 'Horn & Electrical'),
        ('ኣቸራተር', 'Accelerator', 7, 'Engine Controls'),
        ('ኣሞርዛተር', 'Shock Absorber', 6, 'Suspension'),
        ('ዓብይ ሪለይ', 'Large Relay', 4, 'Electrical & Relays'),
        ('3555 ኩችኔት ሲል/seal/', 'Bearing Seal 3555', 6, 'Bearings & Seals'),
        ('248080 ኩችኔት ሲል/seal/', 'Bearing Seal 248080', 7, 'Bearings & Seals'),
        ('ሞሶ', 'Spring', 0, 'Suspension'),
        ('ቤሪንግ ዓብይ 355520', 'Large Bearing 355520', 13, 'Bearings'),
        ('መሪ ቤሪንግ ናይ ታሕቲ', 'Lower Steering Bearing', 13, 'Steering & Bearings'),
        ('6304 RS ቤሪንግ', '6304 RS Bearing', 13, 'Bearings'),
        ('6006 ቤሪንግ', '6006 Bearing', 13, 'Bearings'),
        ('ዋንጫ መሪ ቤሪንግ ናይ ላዕሊ', 'Upper Steering Bearing Cup', 41, 'Steering & Bearings'),
        ('ፍሬን ሻራ ናይ ቅድሚት', 'Front Brake Shoe', 8, 'Braking System'),
        ('ፒስተንቺኒ ናይ ቅድሚት', 'Front Brake Wheel Cylinder', 28, 'Braking System'),
        ('እስሚያስ ናይ ቅድሚት', 'Front Half Shaft', 13, 'Drivetrain & Axles'),
        ('ሸራ ናይ ድሕሪት', 'Rear Brake Shoe', 52, 'Braking System'),
        ('complete ናይ ድሕሪት ፍሬን ሸራ', 'Complete Rear Brake Shoe Set', 4, 'Braking System'),
        ('ስፕሪንግ ናይ ፍሬን ሸራ', 'Brake Shoe Spring', 27, 'Braking System'),
        ('ጎሚኒ ᣝሞርዛቶር', 'Shock Absorber Rubber', 11, 'Suspension'),
        ('Radio', 'Radio', 11, 'Audio & Electronics'),
        ('wiper ጎማ', 'Wiper Rubber', 7, 'Wipers & Exterior'),
        ('ኳድሮ ቁልፊ', 'Quadro Key / Ignition Switch Key Set', 4, 'Ignition & Electrical'),
        ('ስፒከር', 'Speaker', 11, 'Audio & Electronics'),
        ('እጂ ፍሬን ካቦ', 'Hand Brake Cable', 4, 'Braking System'),
        ('እግሪ ፍሬን ካቦ', 'Foot Brake Cable', 4, 'Braking System'),
        ('ብሮንዝ ሃይድሮሊክ', 'Hydraulic Bronze Bushing', 4, 'Hydraulics & Bushings'),
        ('ብሮንዝ ናይ ቅድሚት', 'Front Bronze Bushing', 7, 'Suspension & Bushings'),
        ('ብሬክ ናይ ድሕሪት', 'Rear Brake Assembly', 4, 'Braking System'),
        ('ኣክዝል', 'Axle', 3, 'Drivetrain & Axles'),
        ('ኣንግል ሆዝ', 'Angle Hose', 4, 'Hoses & Hydraulics'),
        ('Complete ናይ ቅድሚት ፍሬን ሸራ', 'Complete Front Brake Shoe Set', 4, 'Braking System'),
        ('ባውዛ ብፅምዲ', 'Pair of Bushings', 4, 'Suspension & Bushings'),
        ('ሙሉእ እጅ ፍሬን ብረት', 'Complete Hand Brake Metal Assembly', 6, 'Braking System'),
        ('ፍሬቻ ናይ ድሕሪት', 'Rear Turn Signal', 7, 'Lighting & Signals'),
        ('ማስተር ሲሊንደር', 'Master Cylinder', 4, 'Braking System'),
        ('ስፖክዮ', 'Side Mirror', 4, 'Mirrors & Exterior'),
        ('ዲስፕለይ /Dispay/', 'Display Unit', 4, 'Electronics'),
        ('ፍዩዝ 20A', 'Fuse 20A', 29, 'Fuses & Electrical'),
        ('ፒስተንቺኒ ናይ ድሕሪት', 'Rear Brake Wheel Cylinder', 13, 'Braking System'),
        ('ፍዩዝ 5A', 'Fuse 5A', 28, 'Fuses & Electrical'),
        ('ፍዩዝ 10A', 'Fuse 10A', 28, 'Fuses & Electrical'),
    ]
    try:
        existing = {p.name for p in SparePart.query.all()}
        branch = Branch.query.first()
        if not branch:
            return jsonify({'message': 'No branch found'}), 400
        added = 0
        for name_tigrinya, name, qty, category in PARTS:
            if name in existing:
                continue
            p = SparePart(
                part_number=f"SP-{int(time.time() * 1000) % 1000000:06d}",
                name=name, name_tigrinya=name_tigrinya, category=category,
                quantity=qty, unit_price=2000, cost_price=2000, branch_id=branch.id,
            )
            db.session.add(p)
            added += 1
        db.session.commit()
        return jsonify({'message': f'Imported {added} spare parts ({len(PARTS) - added} already existed)'}), 200
    except Exception as e:
        db.session.rollback()
        logger.error(f'Import spare parts failed: {e}', exc_info=True)
        return jsonify({'message': 'Import failed. Please try again.'}), 500
