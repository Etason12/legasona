import os
import sys
sys.path.insert(0, os.path.dirname(__file__))

from app import create_app, db
from app.models import Vehicle, Branch

app = create_app()

EXCEL_PATH = os.path.abspath(os.path.join(os.path.dirname(__file__), '../attachment.xlsx'))

def import_vehicles():
    import openpyxl
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    col_map = {
        'Serial': 'vin',
        'Model': 'model',
        'Motor/Engine No': 'engine_number',
        'Price': 'selling_price',
        'cost': 'cost_price',
    }

    col_idx = {}
    for label, key in col_map.items():
        try:
            col_idx[key] = headers.index(label)
        except ValueError:
            print(f'ERROR: column "{label}" not found in Excel. Headers: {headers}')
            sys.exit(1)

    mekelle = Branch.query.filter_by(name='Mekelle').first()
    if not mekelle:
        print('ERROR: Mekelle branch not found. Run seeding first.')
        sys.exit(1)

    imported = 0
    skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        vin = str(row[col_idx['vin']]).strip()
        if not vin:
            continue
        if Vehicle.query.filter_by(vin=vin).first():
            print(f'SKIP (exists): {vin}')
            skipped += 1
            continue

        model = str(row[col_idx['model']]).strip() if row[col_idx['model']] else ''
        engine = str(row[col_idx['engine_number']]).strip() if row[col_idx['engine_number']] else ''
        raw_price = row[col_idx['selling_price']]
        raw_cost = row[col_idx['cost_price']]
        selling_price = float(raw_price) if raw_price else 0.0
        cost_price = float(raw_cost) if raw_cost else 0.0

        v = Vehicle(
            vin=vin,
            model=model,
            engine_number=engine,
            selling_price=selling_price,
            cost_price=cost_price,
            type='3-wheel',
            power_type='electric',
            color='blue',
            branch_id=mekelle.id,
            status='available',
        )
        db.session.add(v)
        imported += 1

    db.session.commit()
    print(f'\nDone. Imported: {imported}, Skipped (already exist): {skipped}')

if __name__ == '__main__':
    with app.app_context():
        import_vehicles()
